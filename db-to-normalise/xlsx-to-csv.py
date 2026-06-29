#!/usr/bin/env python3
"""
One-time converter: export-account-retail.pro.xlsx  ->  export-account-retail.csv

The xlsx is the *no-email* retail cohort (customers with email are already in the
CRM). It is richer than the legacy CSV (stable Cust SID, separate First/Last,
sales value, store credit) but has three traps this converter neutralises so the
browser pipeline can ingest it as plain text:

  1. Cust SID is stored as a QUOTED signed integer  ("-4959087493461372932")
     -> we strip the surrounding quotes and keep it as text (never a number).
  2. Dates are a MIX of DD/MM/YYYY strings and Excel date serials. Excel mangled
     the serial ones (day/month ambiguous). We emit DD/MM/YYYY text for both and
     set `date_excel_serial=1` on the rows whose dates came from a serial, so
     downstream knows their day/month may be swapped.
  3. A few hundred e-mails were typed into the Phone1/Phone2 columns. We salvage
     them into the `email` column and blank the phone instead.

Output: semicolon-delimited, UTF-8 with BOM, header names matching the retail
parser's HEADER_MAP (src/lib/import/retail-parse.ts).
"""
import csv
import datetime
import sys

import openpyxl

SRC = "export-account-retail.pro.xlsx"
OUT = "export-account-retail.csv"

# 1-based column indices in the source sheet.
COL = {
    "last": 1, "first": 2, "first_sale": 3, "last_sale": 4, "store_credit": 5,
    "cust_sid": 6, "country": 7, "country_code": 8, "create_dt": 10,
    "store_name": 12, "email": 13, "sub": 14, "str_code": 15,
    "total_sales": 16, "total_units": 17, "ytd_sales": 18,
    "phone1": 19, "phone2": 20, "cust_id": 21, "compl1": 22, "compl2": 23,
}

OUT_HEADER = [
    "Cust SID", "Cust ID", "First", "Last", "Store Name", "Sub", "Country Code",
    "Email", "Phone1", "Phone2", "Compl 1", "Compl 2",
    "First Sale Dt", "Last Sale Dt", "Total Sales", "Total Units", "YTD Sales",
    "Store Credit", "date_excel_serial",
]


def fmt_date(v):
    """Return (DD/MM/YYYY string, came_from_serial)."""
    if v is None:
        return "", False
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.day:02d}/{v.month:02d}/{v.year:04d}", True
    return str(v).strip(), False


def as_text(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    # Some source cells (e.g. multi-word store names, Cust SID) are wrapped in
    # literal double quotes — strip one surrounding pair for clean text.
    if len(s) >= 2 and s[0] == '"' and s[-1] == '"':
        s = s[1:-1].strip()
    return s


def looks_email(v):
    return "@" in v


def main():
    wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    n = 0
    salvaged_emails = 0
    serial_dates = 0
    seen_sid = set()
    dup_sid = 0

    with open(OUT, "w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(OUT_HEADER)

        for r in ws.iter_rows(min_row=2, values_only=True):
            def g(key):
                return r[COL[key] - 1]

            sid = as_text(g("cust_sid")).strip('"').strip()
            if not sid:
                continue  # SID is the key; a row without it is unusable
            if sid in seen_sid:
                dup_sid += 1
                continue
            seen_sid.add(sid)

            email = as_text(g("email"))
            phone1 = as_text(g("phone1"))
            phone2 = as_text(g("phone2"))

            # Salvage e-mails mis-filed in the phone columns.
            if not looks_email(email):
                if looks_email(phone1):
                    email, phone1 = phone1, ""
                    salvaged_emails += 1
                elif looks_email(phone2):
                    email, phone2 = phone2, ""
                    salvaged_emails += 1
            # A phone that still contains '@' is junk — drop it.
            if looks_email(phone1):
                phone1 = ""
            if looks_email(phone2):
                phone2 = ""

            first_sale, s1 = fmt_date(g("first_sale"))
            last_sale, s2 = fmt_date(g("last_sale"))
            compl1, s3 = fmt_date(g("compl1"))
            compl2, s4 = fmt_date(g("compl2"))
            from_serial = any([s1, s2, s3, s4])
            if from_serial:
                serial_dates += 1

            w.writerow([
                sid,
                as_text(g("cust_id")),
                as_text(g("first")),
                as_text(g("last")),
                as_text(g("store_name")),
                as_text(g("sub")),
                as_text(g("country_code")),
                email,
                phone1,
                phone2,
                compl1,
                compl2,
                first_sale,
                last_sale,
                as_text(g("total_sales")),
                as_text(g("total_units")),
                as_text(g("ytd_sales")),
                as_text(g("store_credit")),
                "1" if from_serial else "0",
            ])
            n += 1

    print(f"Wrote {OUT}: {n} rows", file=sys.stderr)
    print(f"  unique Cust SID: {len(seen_sid)} (dupes skipped: {dup_sid})", file=sys.stderr)
    print(f"  e-mails salvaged from phone columns: {salvaged_emails}", file=sys.stderr)
    print(f"  rows with Excel-serial (ambiguous) dates: {serial_dates}", file=sys.stderr)


if __name__ == "__main__":
    main()
